#!/usr/bin/env python3
import pdfplumber
import fitz
import csv
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Column x-boundaries (from word coordinate analysis of the PDF layout)
X_DATE_END  = 116
X_FNAME_END = 236
X_DRONE_END = 361
X_DUR_END   = 412
# Location: x0 >= 412

ROW_TOL = 1.5   # pixels to bucket words into the same visual row

PAGE_HEADER_MAX_TOP = 60   # repeating page header occupies top ~60px of every page
TCPDF_MIN_TOP_RATIO = 0.99 # TCPDF watermark sits in the bottom 1% of page height

FIELDS = [
    'page',
    'date', 'time_local', 'time_utc', 'flight_name', 'drone', 'duration',
    'location_address', 'lat', 'lon',
    'landing_time_utc', 'flight_type',
    'project_reference', 'external_case_id',
    'operation_type', 'legal_rule', 'sunset_sunrise',
    'personnel', 'personnel_notes', 'equipment_onboard',
    'nb_landing', 'distance', 'max_altitude',
    'cloud_pct', 'visibility', 'temperature', 'wind', 'humidity_pct',
    'pressure', 'precipitation',
    'igc_url', 'kml_url',
    'notes',
]


# ---------------------------------------------------------------------------
# Row helpers
# ---------------------------------------------------------------------------

def group_by_row(words, tol=ROW_TOL):
    buckets = {}
    for w in words:
        key = round(w['top'] / tol)
        buckets.setdefault(key, []).append(w)
    return [sorted(v, key=lambda w: w['x0']) for _, v in sorted(buckets.items())]


def row_text(row):
    return ' '.join(w['text'] for w in row)


def col_text(row, x_min, x_max):
    return ' '.join(w['text'] for w in row if x_min <= w['x0'] < x_max).strip()


def is_col_header(row):
    t = row_text(row)
    return 'Date' in t and 'Flight' in t and 'Duration' in t and 'Location' in t


# ---------------------------------------------------------------------------
# Link extraction  (pymupdf — preserves hyperlink URIs)
# ---------------------------------------------------------------------------

def build_link_map(pdf_path):
    """Return dict: (page_num, y_bucket) -> {igc_url, kml_url}"""
    link_map = {}
    doc = fitz.open(pdf_path)
    for pg_num, page in enumerate(doc):
        for lk in page.get_links():
            uri = lk.get('uri', '')
            if not uri:
                continue
            # y-bucket at ROW_TOL resolution
            y_key = round(lk['from'].y0 / ROW_TOL)
            key = (pg_num, y_key)
            entry = link_map.setdefault(key, {})
            if 'getIGCData' in uri:
                entry['igc_url'] = uri
            elif 'getKMLData' in uri:
                entry['kml_url'] = uri
    doc.close()
    return link_map


# ---------------------------------------------------------------------------
# PDF text extraction  (pdfplumber — good word coordinates)
# Each row is tagged with (page_num, top) for link lookup
# ---------------------------------------------------------------------------

def extract_rows(pdf_path):
    """Return list of (page_num, top_coord, [word, ...]) tuples."""
    all_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for pg_num, page in enumerate(pdf.pages):
            page_height = page.height
            words = page.extract_words()
            for row in group_by_row(words):
                if not row:
                    continue
                top = row[0]['top']
                if top < PAGE_HEADER_MAX_TOP:              # repeating page header
                    continue
                if top > page_height * TCPDF_MIN_TOP_RATIO:  # TCPDF watermark line
                    continue
                t = row_text(row)
                if re.match(r'^\d+ / \d+$', t):           # page-number footer
                    continue
                if re.match(r'\d+ Flights filtered', t):   # report summary line
                    continue
                all_rows.append((pg_num, top, row))
    return all_rows


# ---------------------------------------------------------------------------
# Metadata
# ---------------------------------------------------------------------------

def parse_metadata(all_rows, pdf_path):
    meta = {}
    for _, _, row in all_rows[:20]:
        t = row_text(row)
        m = re.search(r'Period:\s*(\S+)\s+to\s+(\S+)', t)
        if m:
            meta['period_start'], meta['period_end'] = m.group(1), m.group(2)
        m = re.search(r'Name:\s*(.+?)\s+Email:\s*(\S+)', t)
        if m:
            meta['name'], meta['email'] = m.group(1), m.group(2)
        m = re.search(r'FLIGHTS\s*\((\d+)', t)
        if m:
            meta['total_flights'] = m.group(1)
        m = re.search(r'Flying time.*?:\s*(\S+)', t)
        if m:
            meta['flying_time'] = m.group(1)

    # First page top header: agency (right column) and export date (centre column)
    with pdfplumber.open(pdf_path) as pdf:
        first_words = [w for w in pdf.pages[0].extract_words() if w['top'] < 45]
        top_rows = {}
        for w in first_words:
            key = round(w['top'] / ROW_TOL)
            top_rows.setdefault(key, []).append(w)
        for words in top_rows.values():
            words.sort(key=lambda w: w['x0'])
            line = ' '.join(w['text'] for w in words)
            if re.search(r'\d+\s+\w+\s+\d{4}', line):
                # "10 March 2026" — strip any leading agency text on same row
                meta['export_date'] = re.search(r'\d+\s+\w+\s+\d{4}', line).group(0)
            if 'Police' in line or 'Department' in line:
                # Agency name is the right-column words on the top row
                agency_words = [w['text'] for w in words if w['x0'] > 400]
                if agency_words:
                    meta['agency'] = ' '.join(agency_words)

        # Last page footer: "1279 Flights filtered / Total Flying Time: 300:48:52"
        last_text = pdf.pages[-1].extract_text() or ''
    m = re.search(r'(\d+)\s+Flights filtered\s*/\s*Total Flying Time:\s*(\S+)', last_text)
    if m:
        meta['filtered_flights'] = m.group(1)
        meta['filtered_flying_time'] = m.group(2)

    return meta


# ---------------------------------------------------------------------------
# Columnar section parser
# ---------------------------------------------------------------------------

def parse_col_rows(col_rows):
    date_parts, fname_parts, drone_parts, dur_parts, loc_parts = [], [], [], [], []
    for _, _, row in col_rows:
        d  = col_text(row, 0,           X_DATE_END)
        f  = col_text(row, X_DATE_END,  X_FNAME_END)
        dr = col_text(row, X_FNAME_END, X_DRONE_END)
        du = col_text(row, X_DRONE_END, X_DUR_END)
        lo = col_text(row, X_DUR_END,   9999)
        if d:  date_parts.append(d)
        if f:  fname_parts.append(f)
        if dr: drone_parts.append(dr)
        if du: dur_parts.append(du)
        if lo: loc_parts.append(lo)

    date_str = time_local = time_utc = ''
    for p in date_parts:
        if re.match(r'\d{4}-\d{2}-\d{2}', p):
            date_str = p
        elif 'UTC' in p:
            time_utc = re.sub(r'\s*UTC$', '', p).strip()
        elif re.match(r'\d{2}:\d{2}:\d{2}', p):
            time_local = p

    loc_full = ' '.join(loc_parts)
    m = re.search(r'\(([0-9.-]+),\s*([0-9.-]+)\)', loc_full)
    lat = m.group(1) if m else ''
    lon = m.group(2) if m else ''
    loc_addr = re.sub(r'\s*\([^)]+\)', '', loc_full).strip()

    return {
        'date': date_str, 'time_local': time_local, 'time_utc': time_utc,
        'flight_name': ' '.join(fname_parts),
        'drone': ' '.join(drone_parts),
        'duration': ' '.join(dur_parts),
        'location_address': loc_addr,
        'lat': lat, 'lon': lon,
    }


# ---------------------------------------------------------------------------
# Labeled-section parser
# ---------------------------------------------------------------------------

def parse_labeled(text):
    f = {}

    m = re.search(r'Landing Time:(\S+)\s+UTC\s+Flight Type:[ \t]*([^\n]*)', text)
    if m:
        f['landing_time_utc'] = m.group(1)
        f['flight_type'] = m.group(2).strip()

    # "External Case\nID:" or "External Case ID:" may appear at end of Project line
    m = re.search(
        r'Project/Job Reference:\s*(.*?)(?:\s+External Case\s*\n?\s*ID:\s*(\S+))?(?=\n|$)',
        text, re.MULTILINE
    )
    if m:
        f['project_reference'] = m.group(1).strip()
        f['external_case_id']  = m.group(2).strip() if m.group(2) else ''
    else:
        f['project_reference'] = f['external_case_id'] = ''

    m = re.search(r'Operation Type:\s*(.*?)\s+Legal Rule:\s*(.*?)\s+Sunset / Sunrise:[ \t]*(\S*)', text)
    if m:
        f['operation_type'] = m.group(1).strip()
        f['legal_rule']     = m.group(2).strip()
        f['sunset_sunrise'] = m.group(3).strip()

    m = re.search(r'Personnel:\s*([^\n]*?)\s+Personnel notes:[ \t]*([^\n]*)', text)
    if m:
        f['personnel']       = m.group(1).strip()
        f['personnel_notes'] = m.group(2).strip()

    m = re.search(r'Equipment onboard:\s*(.*?)$', text, re.MULTILINE)
    f['equipment_onboard'] = m.group(1).strip() if m else ''

    m = re.search(r'Nb landing:\s*(\S+)\s+Distance:\s*(.*?)\s+Max altitude:\s*(.*?)$', text, re.MULTILINE)
    if m:
        f['nb_landing']   = m.group(1)
        f['distance']     = m.group(2).strip()
        f['max_altitude'] = m.group(3).strip()

    # Cloud/Humidity line may wrap so "% Pressure" starts the next line
    m = re.search(
        r'Cloud:\s*(\d*)\s*%?\s*Visibility:\s*(.*?)\s+Temperature:\s*(.*?)\s+Wind:\s*(.*?)\s+Humidity:\s*(\d*)',
        text
    )
    if m:
        f['cloud_pct']   = m.group(1)
        f['visibility']  = m.group(2).strip()
        f['temperature'] = m.group(3).strip()
        f['wind']        = m.group(4).strip()
        f['humidity_pct']= m.group(5)

    m = re.search(r'Pressure:\s*([\d.]*)\s+Precipitation:[ \t]*(.*?)$', text, re.MULTILINE)
    if m:
        f['pressure']      = m.group(1)
        f['precipitation'] = m.group(2).strip()

    # Notes: strip trailing "IGC File KML File" label row if present
    m = re.search(r'Notes:\s*(.*)', text, re.DOTALL)
    notes = m.group(1).strip() if m else ''
    notes = re.sub(r'\s*(IGC File\s*KML File|IGC File|KML File)\s*$', '', notes).strip()
    f['notes'] = notes

    return f


# ---------------------------------------------------------------------------
# Flight parser
# ---------------------------------------------------------------------------

def parse_flights(all_rows, link_map):
    flights = []
    i, n = 0, len(all_rows)

    while i < n:
        pg_num, top, row = all_rows[i]
        if is_col_header(row):
            i += 1  # skip column-header row

            # Columnar data rows — up to "Landing Time:" or next col header
            col_rows = []
            while i < n:
                pg, tp, rw = all_rows[i]
                t = row_text(rw)
                if t.startswith('Landing Time:') or is_col_header(rw):
                    break
                col_rows.append((pg, tp, rw))
                i += 1

            # Labeled rows — up to next col header
            igc_url = kml_url = ''
            labeled_lines = []
            while i < n:
                pg, tp, rw = all_rows[i]
                if is_col_header(rw):
                    break
                t = row_text(rw)
                labeled_lines.append(t)

                # Check if this row has hyperlinks attached (±1 y_key for library rounding diff)
                y_key = round(tp / ROW_TOL)
                entry = {}
                for delta in (0, 1, -1):
                    entry = link_map.get((pg, y_key + delta), {})
                    if entry:
                        break
                if entry.get('igc_url'):
                    igc_url = entry['igc_url']
                if entry.get('kml_url'):
                    kml_url = entry['kml_url']

                i += 1

            flight = parse_col_rows(col_rows)
            flight['page'] = (col_rows[0][0] + 1) if col_rows else (pg_num + 1)
            flight.update(parse_labeled('\n'.join(labeled_lines)))
            flight['igc_url'] = igc_url
            flight['kml_url'] = kml_url
            flights.append(flight)
        else:
            i += 1

    return flights


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    include_metadata = '--comments' in args
    args = [a for a in args if a != '--comments']

    if not args:
        print(f"Usage: {sys.argv[0]} [--comments] <input.pdf> [output.csv]", file=sys.stderr)
        sys.exit(1)

    pdf_path = args[0]
    out_path = args[1] if len(args) > 1 else str(Path(pdf_path).with_suffix('.csv'))

    print("Building link map…")
    link_map = build_link_map(pdf_path)

    print("Extracting rows…")
    all_rows = extract_rows(pdf_path)

    print("Parsing metadata…")
    meta = parse_metadata(all_rows, pdf_path)

    print("Parsing flights…")
    flights = parse_flights(all_rows, link_map)

    metadata = [
        ("Operations Report", ""),
        ("Agency",            meta.get('agency', '')),
        ("Name",              meta.get('name', '')),
        ("Email",             meta.get('email', '')),
        ("Export date",       meta.get('export_date', '')),
        ("Period",            f"{meta.get('period_start','')} to {meta.get('period_end','')}"),
        ("Flights",           meta.get('total_flights', '')),
        ("Filtered flights",  meta.get('filtered_flights', '')),
        ("Flying time",       meta.get('flying_time', '')),
    ]

    print()
    for label, value in metadata:
        print(f"  {label + ':':20s} {value}".rstrip())
    print(f"  {'Parsed flights:':20s} {len(flights)}")
    print()

    def clean(v):
        return re.sub(r'[\r\n]+', ' ', str(v)).strip()

    with open(out_path, 'w', newline='', encoding='utf-8-sig') as fh:
        if include_metadata:
            lines = [f"# {label}: {value}" if value else f"# {label}" for label, value in metadata]
            fh.write('\r\n'.join(lines) + '\r\n')

        writer = csv.DictWriter(fh, fieldnames=FIELDS, extrasaction='ignore', dialect='excel', quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for fl in flights:
            writer.writerow({k: clean(fl.get(k, '')) for k in FIELDS})

    print(f"Written {out_path}")

    xlsx_path = str(Path(out_path).with_suffix('.xlsx'))
    write_xlsx(xlsx_path, metadata, flights, clean)
    print(f"Written {xlsx_path}")


def write_xlsx(xlsx_path, metadata, flights, clean):
    wb = Workbook()

    # --- Flights sheet (first/active) ---
    fs = wb.active
    fs.title = "Flights"
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")

    fs.append(FIELDS)
    for cell in fs[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    fs.freeze_panes = "A2"

    for fl in flights:
        fs.append([clean(fl.get(k, '')) for k in FIELDS])

    # Auto-width: cap at 60 chars
    for col_idx, field in enumerate(FIELDS, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(field),
            max((len(str(clean(fl.get(field, '')))) for fl in flights), default=0)
        )
        fs.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # --- Metadata sheet (second) ---
    ms = wb.create_sheet("Metadata")
    label_font = Font(bold=True)
    for label, value in metadata:
        ms.append([label, value])
        ms.cell(ms.max_row, 1).font = label_font
    ms.column_dimensions['A'].width = 22
    ms.column_dimensions['B'].width = 50

    wb.save(xlsx_path)


if __name__ == '__main__':
    main()